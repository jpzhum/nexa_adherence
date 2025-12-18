
# -*- coding: utf-8 -*-
"""
Atualiza (ou cria) o banco Excel 'bd BI.xlsx' com 4 abas para o Power BI:
- Tbl_Turnos
- Tbl_comparativo
- Comparativo1
- Base de Dados

Integra com data_service.base_dir_dados() para resolver o diretório raiz:
C:\Sistema de Análise 2.0\TransicaoPY_Dados (Windows), ou conforme config/env.

Uso:
    from interface.services.exportbi_service import atualizar_bd_excel
    caminho = atualizar_bd_excel(final_df, resumos, out_dir=r"C:\Sistema de Análise 2.0\TransicaoPY_Dados", filename="bd BI.xlsx")
"""

import os
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Serviços da sua app
from interface.services.data_service import base_dir_dados
from interface.services.config_service import get_config


# ===================== FORMATAÇÃO OPENPYXL =====================

def _formatar_ws(ws):
    """Aplica AutoFilter, Freeze Panes e ajusta largura de colunas."""
    try:
        if ws.max_row >= 2 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = 'A2'
        for col in ws.columns:
            max_len = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    parts = str(cell.value).split('\n')
                    max_len = max(max_len, max(len(p) for p in parts))
            ws.column_dimensions[letter].width = min(max_len + 2, 60)
    except Exception:
        # Não interromper fluxo por causa de formatação
        pass

def _formatar_percentuais(ws):
    """
    Formata colunas percentuais:
    - Se o header contém '%', aplica formato '0.00%'.
    - Se o valor estiver > 1, divide por 100 (Excel usa escala 0–1).
    """
    try:
        header_values = [ws.cell(row=1, column=c).value or '' for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                head = str(header_values[c - 1])
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)) and '%' in head:
                    val = cell.value
                    if val is not None:
                        cell.value = (val / 100.0) if val > 1 else val
                    cell.number_format = '0.00%'
    except Exception:
        pass

def _aplicar_formatacao_pos_escrita(xlsx_path: Path, sheet_names: list):
    """
    Reabre o arquivo e aplica as formatações nas abas especificadas.
    """
    try:
        wb = load_workbook(xlsx_path)
        for sheet_name in sheet_names:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                _formatar_ws(ws)
                _formatar_percentuais(ws)
        wb.save(xlsx_path)
    except Exception:
        # Se não conseguir formatar, mantém sem formatação
        pass


# ===================== HELPERS DE CAMINHO =====================

def _obter_caminho_bd_excel(out_dir: str = None, filename: str = None) -> Path:
    """
    Resolve o caminho final do 'bd BI.xlsx' priorizando:
    1) argumentos explícitos (out_dir/filename),
    2) config (powerbi_export_dir/powerbi_export_file),
    3) base_dir_dados() (data_service),
    4) fallback: ~/TransicaoPY_Dados.
    """
    # 1) Argumentos explícitos
    if out_dir:
        base = Path(out_dir)
        base.mkdir(parents=True, exist_ok=True)
        fname = filename or "bd BI.xlsx"
        return base / fname

    # 2) Config
    cfg = {}
    try:
        cfg = get_config() or {}
    except Exception:
        cfg = {}
    export_dir = cfg.get('powerbi_export_dir')
    export_file = cfg.get('powerbi_export_file') or "bd BI.xlsx"
    if export_dir:
        pdir = Path(export_dir)
        if pdir.exists() or _safe_mkdir(pdir):
            return pdir / export_file

    # 3) data_service
    base = Path(base_dir_dados())
    base.mkdir(parents=True, exist_ok=True)
    return base / export_file

def _safe_mkdir(p: Path) -> bool:
    try:
        p.mkdir(parents=True, exist_ok=True)
        return True
    except Exception:
        return False


# ===================== BUILDERS DAS 4 TABELAS =====================

def _qtd_turnos_esperados(row):
    return 1 if str(row.get('Escala', '')).upper().strip() == 'ADM' else 3

def build_tbl_turnos(final: pd.DataFrame) -> pd.DataFrame:
    df = final.copy()
    for col in ['TURNO A', 'TURNO B', 'TURNO C', 'Escala']:
        if col not in df.columns:
            df[col] = '-' if col.startswith('TURNO') else 'PADRÃO'

    long = pd.melt(
        df[['Data Cabeçalho', 'Equipamento', 'Escala', 'TURNO A', 'TURNO B', 'TURNO C']],
        id_vars=['Data Cabeçalho', 'Equipamento', 'Escala'],
        value_vars=['TURNO A', 'TURNO B', 'TURNO C'],
        var_name='Turno', value_name='Flag'
    )
    long['Considerar'] = long.apply(
        lambda r: (r['Turno'] == 'TURNO A') if str(r['Escala']).upper().strip() == 'ADM' else True,
        axis=1
    )
    long = long[long['Considerar']]

    resumo = long.groupby('Turno').agg(
        Entregues=('Flag', lambda s: (s == 'OK').sum()),
        Nao_entregues=('Flag', lambda s: (s != 'OK').sum())
    ).reset_index()

    resumo['% Entregues'] = (
        resumo['Entregues'] / (resumo['Entregues'] + resumo['Nao_entregues']).replace(0, pd.NA)
    ) * 100
    resumo['% Não entregues'] = 100 - resumo['% Entregues']
    resumo = resumo.rename(columns={'Nao_entregues': 'Não entregues'})
    resumo['% Entregues'] = resumo['% Entregues'].round(2)
    resumo['% Não entregues'] = resumo['% Não entregues'].round(2)

    return resumo[['Turno', 'Entregues', 'Não entregues', '% Entregues', '% Não entregues']]

def build_tbl_comparativo(final: pd.DataFrame, meta_percentual: float = 85.0) -> pd.DataFrame:
    df = final.copy()
    for col in ['Gestor', 'Agrup Equipamento', 'TURNO A', 'TURNO B', 'TURNO C', '% Aderência', 'Status', 'Escala']:
        if col not in df.columns:
            df[col] = 'Não Informado' if col in ['Gestor', 'Agrup Equipamento', 'Status'] \
                      else ('-' if col.startswith('TURNO') else 0)

    grp = df.groupby(['Gestor', 'Agrup Equipamento']).agg(
        Entregues=('Entregues', 'sum'),
        Faltantes=('Faltantes', 'sum'),
        Aderencia_media=('% Aderência', 'mean'),
        TurnoA=('TURNO A', lambda s: (s == 'OK').sum()),
        TurnoB=('TURNO B', lambda s: (s == 'OK').sum()),
        TurnoC=('TURNO C', lambda s: (s == 'OK').sum()),
    ).reset_index()

    grp['Não entregues'] = grp['Faltantes']
    grp['[%] Entregues'] = grp['Entregues'] / (grp['Entregues'] + grp['Não entregues']).replace(0, pd.NA) * 100
    grp['[%] Entregues'] = grp['[%] Entregues'].round(2)
    grp['% De entregas'] = grp['Aderencia_media'].round(2)
    grp['Linha de % da meta'] = meta_percentual

    out = grp.rename(columns={
        'Agrup Equipamento': 'Agrupamento',
        'Gestor': 'Gestão',
        'TurnoA': 'Turno A',
        'TurnoB': 'Turno B',
        'TurnoC': 'Turno C'
    })

    cols = [
        'Agrupamento', 'Gestão',
        'Entregues', 'Não entregues',
        'Turno A', 'Turno B', 'Turno C',
        '[%] Entregues', '% De entregas', 'Linha de % da meta'
    ]
    return out[cols]

def build_comparativo1(final: pd.DataFrame) -> pd.DataFrame:
    df = final.copy()
    df['Data'] = pd.to_datetime(df['Data Cabeçalho'], errors='coerce')
    df['Mês'] = df['Data'].dt.to_period('M').astype(str)
    df['Frota'] = df['Equipamento']
    df['Qtd Turnos'] = df.apply(_qtd_turnos_esperados, axis=1)
    df['Porcentagem'] = df['% Aderência']
    df['Apontamentos Faltantes'] = df['Faltantes']
    for col in ['Tipo', 'Subprocesso', 'Líder 1', 'Líder 2', 'Líder 3',
                'Insights_HTML', 'Insights_HTML_Turnos', 'KPI_Evolucao_HTML',
                'Linha_Aumento', 'Linha_Queda']:
        if col not in df.columns:
            df[col] = ''
    cols = [
        'Data', 'Mês', 'Frota', 'Escala',
        'TURNO A', 'TURNO B', 'TURNO C',
        'Qtd Turnos', 'Porcentagem', 'Apontamentos Faltantes',
        'Status', 'Tipo', 'Subprocesso',
        'Líder 1', 'Líder 2', 'Líder 3',
        'Insights_HTML', 'Insights_HTML_Turnos', 'KPI_Evolucao_HTML',
        'Agrup Equipamento', 'Gestor'
    ]
    return df[cols]

def build_base_de_dados(final: pd.DataFrame) -> pd.DataFrame:
    df = final.copy()
    df['Data Cabeçalho'] = pd.to_datetime(df['Data Cabeçalho'], errors='coerce').dt.strftime('%d/%m/%Y')
    cols = [
        'Data Cabeçalho', 'Equipamento', 'Agrup Equipamento', 'Gestor',
        'Escala', 'Entregues', 'Faltantes', '% Aderência', 'Status',
        'TURNO A', 'TURNO B', 'TURNO C'
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = '' if c in ['Data Cabeçalho', 'Equipamento', 'Agrup Equipamento', 'Gestor', 'Escala', 'Status'] else 0
    return df[cols]


# ===================== ESCRITA DAS ABAS VIA PANDAS =====================

def _escrever_dfs(path: Path, dfs_por_aba: dict):
    """
    Escreve/atualiza as abas no arquivo Excel apontado por 'path':
    - Se o arquivo existir: mode='a', if_sheet_exists='replace' (substitui cada aba).
    - Se não existir: mode='w' (cria um novo).
    """
    sheet_names = list(dfs_por_aba.keys())
    file_exists = path.exists()

    # Cria diretório se não existir
    path.parent.mkdir(parents=True, exist_ok=True)

    mode = 'a' if file_exists else 'w'
    with pd.ExcelWriter(path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
        for sheet_name, df in dfs_por_aba.items():
            if not isinstance(df, pd.DataFrame):
                df = pd.DataFrame(df)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Após escrever, aplicar formatações nas abas
    _aplicar_formatacao_pos_escrita(path, sheet_names)


# ===================== API PÚBLICA =====================

def atualizar_bd_excel(
    final_df: pd.DataFrame,
    resumos: dict,
    meta_percentual: float = 85.0,
    out_dir: str = None,
    filename: str = None
) -> str:
    """
    Cria/atualiza o arquivo Excel <out_dir|config|base_dir_dados>/bd BI.xlsx com as 4 abas:
    - Tbl_Turnos
    - Tbl_comparativo
    - Comparativo1
    - Base de Dados

    Retorna o caminho completo do arquivo salvo.
    """
    # Monta as 4 tabelas
    tbl_turnos = build_tbl_turnos(final_df)
    tbl_comp   = build_tbl_comparativo(final_df, meta_percentual=meta_percentual)
    comp1      = build_comparativo1(final_df)
    base       = build_base_de_dados(final_df)

    # Resolve caminho
    path = _obter_caminho_bd_excel(out_dir=out_dir, filename=filename)

    # Escreve abas
    dfs_por_aba = {
        'Tbl_Turnos': tbl_turnos,
        'Tbl_comparativo': tbl_comp,
        'Comparativo1': comp1,
        'Base de Dados': base
    }

    try:
        _escrever_dfs(path, dfs_por_aba)
        return str(path)

    except PermissionError:
        # Arquivo aberto no Excel; salvar alternativo com timestamp
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        alt = path.with_name(f"{path.stem}_alt_{ts}{path.suffix}")
        _escrever_dfs(alt, dfs_por_aba)
        return str(alt)

    except Exception as e:
        # Qualquer outro erro inesperado: tenta salvar em fallback
        fallback = Path(base_dir_dados()) / "bd BI_fallback.xlsx"
        try:
            _escrever_dfs(fallback, dfs_por_aba)
            return str(fallback)
        except Exception:
            # Por fim, relança o erro original para o chamador tratar (UI mostra mensagem)
            raise e

