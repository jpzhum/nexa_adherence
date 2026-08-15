import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SPREADSHEET_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _neutralize_spreadsheet_value(value):
    if not isinstance(value, str):
        return value
    candidate = value.lstrip()
    if candidate == "-":
        return value
    if candidate.startswith(SPREADSHEET_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def neutralize_spreadsheet_formulas(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy whose text cells cannot be interpreted as Excel formulas."""
    safe = df.copy()
    for column in safe.columns:
        safe[column] = safe[column].map(_neutralize_spreadsheet_value)

    if isinstance(safe.index, pd.MultiIndex):
        safe.index = pd.MultiIndex.from_tuples(
            [tuple(_neutralize_spreadsheet_value(value) for value in row) for row in safe.index],
            names=[_neutralize_spreadsheet_value(name) for name in safe.index.names],
        )
    else:
        safe.index = pd.Index(
            [_neutralize_spreadsheet_value(value) for value in safe.index],
            name=_neutralize_spreadsheet_value(safe.index.name),
        )
    return safe


def _formatar_ws(ws):
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                parts = str(cell.value).split("\n")
                max_len = max(max_len, max(len(p) for p in parts))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def _formatar_percentuais(ws):
    header_values = [ws.cell(row=1, column=c).value or "" for c in range(1, ws.max_column + 1)]
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            head = str(header_values[c - 1])
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)) and "%" in head:
                val = cell.value
                if val is not None:
                    cell.value = (val / 100.0) if val > 1 else val
                cell.number_format = "0.00%"


def _normalize_date_column(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    aliases = ["Data Cabecalho", "Data Cabeçalho", "Data Cabe?alho", "Data CabeÃ§alho"]
    for col in aliases:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d/%m/%Y")
            if col != "Data Cabecalho":
                df = df.rename(columns={col: "Data Cabecalho"})
            break
    return df


def exportar_excel(final_df: pd.DataFrame, resumos: dict, caminho: str) -> None:
    df_out = neutralize_spreadsheet_formulas(_normalize_date_column(final_df))

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Base Consolidada", index=False)

        for nome, df_resumo in resumos.items():
            df_sheet = df_resumo
            if hasattr(df_sheet, "to_frame") and not hasattr(df_sheet, "columns"):
                df_sheet = df_sheet.to_frame(name="Valor")
            if not isinstance(df_sheet, pd.DataFrame):
                df_sheet = pd.DataFrame(df_sheet)
            df_sheet = neutralize_spreadsheet_formulas(df_sheet)
            df_sheet.to_excel(
                writer,
                sheet_name=str(nome),
                index=True if df_sheet.index.names is not None else False,
            )

    wb = load_workbook(caminho)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        _formatar_ws(ws)
        _formatar_percentuais(ws)
    wb.save(caminho)
