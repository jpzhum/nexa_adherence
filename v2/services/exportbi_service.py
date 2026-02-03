from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PERCENT_ALIASES: Iterable[str] = (
    "% Aderencia",
    "% Aderencia ",
    "% Aderência",
    "% AderÃªncia",
    "% Ader?ncia",
    "Aderencia",
)


def _formatar_ws(ws) -> None:
    if ws.max_row >= 2 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                parts = str(cell.value).split("\n")
                max_len = max(max_len, max(len(p) for p in parts))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def _formatar_percentuais(ws) -> None:
    header_values = [ws.cell(row=1, column=c).value or "" for c in range(1, ws.max_column + 1)]
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            head = str(header_values[c - 1])
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)) and "%" in head:
                val = cell.value
                cell.value = (val / 100.0) if val > 1 else val
                cell.number_format = "0.00%"


def _aplicar_formatacao_pos_escrita(xlsx_path: Path, sheet_names: list[str]) -> None:
    wb = load_workbook(xlsx_path)
    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _formatar_ws(ws)
            _formatar_percentuais(ws)
    wb.save(xlsx_path)


def _find_percent_column(df: pd.DataFrame) -> Optional[str]:
    for col in PERCENT_ALIASES:
        if col in df.columns:
            return col
    return None


def _normalize_base_columns(final: pd.DataFrame) -> pd.DataFrame:
    df = final.copy()

    if "Data Cabeçalho" in df.columns and "Data Cabecalho" not in df.columns:
        df = df.rename(columns={"Data Cabeçalho": "Data Cabecalho"})
    if "Data Cabe?alho" in df.columns and "Data Cabecalho" not in df.columns:
        df = df.rename(columns={"Data Cabe?alho": "Data Cabecalho"})

    percent_col = _find_percent_column(df)
    if percent_col and percent_col != "% Aderencia":
        df["% Aderencia"] = df[percent_col]
    elif "% Aderencia" not in df.columns:
        df["% Aderencia"] = pd.to_numeric(df.get("Aderencia", 0), errors="coerce").fillna(0)

    defaults = {
        "Data Cabecalho": pd.NaT,
        "Equipamento": "",
        "Agrup Equipamento": "Nao Informado",
        "Gestor": "Nao Informado",
        "Escala": "PADRAO",
        "Status": "Nao Informado",
        "TURNO A": "-",
        "TURNO B": "-",
        "TURNO C": "-",
        "Entregues": 0,
        "Faltantes": 0,
    }
    for key, value in defaults.items():
        if key not in df.columns:
            df[key] = value

    df["Data Cabecalho"] = pd.to_datetime(df["Data Cabecalho"], errors="coerce")
    for numeric_col in ["Entregues", "Faltantes", "% Aderencia"]:
        df[numeric_col] = pd.to_numeric(df[numeric_col], errors="coerce").fillna(0)

    return df


def _qtd_turnos_esperados(row: pd.Series) -> int:
    return 1 if str(row.get("Escala", "")).strip().upper() == "ADM" else 3


def build_tbl_turnos(final: pd.DataFrame) -> pd.DataFrame:
    df = _normalize_base_columns(final)

    long = pd.melt(
        df[["Data Cabecalho", "Equipamento", "Escala", "TURNO A", "TURNO B", "TURNO C"]],
        id_vars=["Data Cabecalho", "Equipamento", "Escala"],
        value_vars=["TURNO A", "TURNO B", "TURNO C"],
        var_name="Turno",
        value_name="Flag",
    )
    long["Considerar"] = long.apply(
        lambda row: (row["Turno"] == "TURNO A")
        if str(row["Escala"]).upper().strip() == "ADM"
        else True,
        axis=1,
    )
    long = long[long["Considerar"]]

    resumo = (
        long.groupby("Turno")
        .agg(
            Entregues=("Flag", lambda s: (s == "OK").sum()),
            nao_entregues=("Flag", lambda s: (s != "OK").sum()),
        )
        .reset_index()
    )

    total = (resumo["Entregues"] + resumo["nao_entregues"]).replace(0, pd.NA)
    resumo["% Entregues"] = (resumo["Entregues"] / total * 100).round(2)
    resumo["% Nao entregues"] = (100 - resumo["% Entregues"]).round(2)
    resumo = resumo.rename(columns={"nao_entregues": "Nao entregues"})

    return resumo[["Turno", "Entregues", "Nao entregues", "% Entregues", "% Nao entregues"]]


def build_tbl_comparativo(final: pd.DataFrame, meta_percentual: float = 85.0) -> pd.DataFrame:
    df = _normalize_base_columns(final)

    grp = (
        df.groupby(["Gestor", "Agrup Equipamento"])
        .agg(
            Entregues=("Entregues", "sum"),
            Faltantes=("Faltantes", "sum"),
            Aderencia_media=("% Aderencia", "mean"),
            TurnoA=("TURNO A", lambda s: (s == "OK").sum()),
            TurnoB=("TURNO B", lambda s: (s == "OK").sum()),
            TurnoC=("TURNO C", lambda s: (s == "OK").sum()),
        )
        .reset_index()
    )

    grp["Nao entregues"] = grp["Faltantes"]
    total = (grp["Entregues"] + grp["Nao entregues"]).replace(0, pd.NA)
    grp["[%] Entregues"] = (grp["Entregues"] / total * 100).round(2)
    grp["% De entregas"] = grp["Aderencia_media"].round(2)
    grp["Linha de % da meta"] = float(meta_percentual)

    out = grp.rename(
        columns={
            "Agrup Equipamento": "Agrupamento",
            "Gestor": "Gestao",
            "TurnoA": "Turno A",
            "TurnoB": "Turno B",
            "TurnoC": "Turno C",
        }
    )

    cols = [
        "Agrupamento",
        "Gestao",
        "Entregues",
        "Nao entregues",
        "Turno A",
        "Turno B",
        "Turno C",
        "[%] Entregues",
        "% De entregas",
        "Linha de % da meta",
    ]
    return out[cols]


def build_comparativo1(final: pd.DataFrame) -> pd.DataFrame:
    df = _normalize_base_columns(final)
    df["Data"] = pd.to_datetime(df["Data Cabecalho"], errors="coerce")
    df["Mes"] = df["Data"].dt.to_period("M").astype(str)
    df["Frota"] = df["Equipamento"]
    df["Qtd Turnos"] = df.apply(_qtd_turnos_esperados, axis=1)
    df["Porcentagem"] = df["% Aderencia"]
    df["Apontamentos Faltantes"] = df["Faltantes"]

    for col in [
        "Tipo",
        "Subprocesso",
        "Lider 1",
        "Lider 2",
        "Lider 3",
        "Insights_HTML",
        "Insights_HTML_Turnos",
        "KPI_Evolucao_HTML",
        "Linha_Aumento",
        "Linha_Queda",
    ]:
        if col not in df.columns:
            df[col] = ""

    cols = [
        "Data",
        "Mes",
        "Frota",
        "Escala",
        "TURNO A",
        "TURNO B",
        "TURNO C",
        "Qtd Turnos",
        "Porcentagem",
        "Apontamentos Faltantes",
        "Status",
        "Tipo",
        "Subprocesso",
        "Lider 1",
        "Lider 2",
        "Lider 3",
        "Insights_HTML",
        "Insights_HTML_Turnos",
        "KPI_Evolucao_HTML",
        "Agrup Equipamento",
        "Gestor",
    ]
    return df[cols]


def build_base_de_dados(final: pd.DataFrame) -> pd.DataFrame:
    df = _normalize_base_columns(final)
    df["Data Cabecalho"] = df["Data Cabecalho"].dt.strftime("%d/%m/%Y")

    cols = [
        "Data Cabecalho",
        "Equipamento",
        "Agrup Equipamento",
        "Gestor",
        "Escala",
        "Entregues",
        "Faltantes",
        "% Aderencia",
        "Status",
        "TURNO A",
        "TURNO B",
        "TURNO C",
    ]
    return df[cols]


def _escrever_dfs(path: Path, dfs_por_aba: Dict[str, pd.DataFrame]) -> None:
    sheet_names = list(dfs_por_aba.keys())
    file_exists = path.exists()

    path.parent.mkdir(parents=True, exist_ok=True)

    mode = "a" if file_exists else "w"
    writer_args = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_args["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(path, **writer_args) as writer:
        for sheet_name, df in dfs_por_aba.items():
            if not isinstance(df, pd.DataFrame):
                df = pd.DataFrame(df)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    _aplicar_formatacao_pos_escrita(path, sheet_names)


def atualizar_bd_excel(
    final_df: pd.DataFrame,
    resumos: dict,
    meta_percentual: float = 85.0,
    out_dir: str | None = None,
    filename: str | None = None,
) -> str:
    _ = resumos
    tbl_turnos = build_tbl_turnos(final_df)
    tbl_comp = build_tbl_comparativo(final_df, meta_percentual=meta_percentual)
    comp1 = build_comparativo1(final_df)
    base = build_base_de_dados(final_df)

    fname = filename or "bd BI.xlsx"
    base_dir = Path(out_dir) if out_dir else Path.cwd()
    path = base_dir / fname

    dfs_por_aba = {
        "Tbl_Turnos": tbl_turnos,
        "Tbl_comparativo": tbl_comp,
        "Comparativo1": comp1,
        "Base de Dados": base,
    }

    try:
        _escrever_dfs(path, dfs_por_aba)
        return str(path)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = path.with_name(f"{path.stem}_alt_{ts}{path.suffix}")
        _escrever_dfs(alt, dfs_por_aba)
        return str(alt)
