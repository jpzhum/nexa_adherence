import pandas as pd
from openpyxl import load_workbook

from v2.services.analysis_service import (
    apply_exclusions,
    normalize_turno_label,
    pivot_turnos,
)
from v2.services.consolidation_service import _dedup_supervisores_by_agrup
from v2.services.dashboard_service import grafico_aderencia_qtd_turno
from v2.services.email_service import montar_html
from v2.services.export_service import exportar_excel
from v2.services.exportbi_service import atualizar_bd_excel, build_base_de_dados
from v2.ui.pages.indicators_template import render_html


def test_apply_exclusions_is_case_insensitive() -> None:
    df = pd.DataFrame(
        {
            "Agrup Equipamento": ["COLHEITA", "plantio"],
            "Equipamento": ["EQ-01", "eq-02"],
        }
    )

    out = apply_exclusions(df, exclusions_agrup=["colheita"], exclusions_frota=["EQ-02"])

    assert out.empty


def test_build_base_de_dados_accepts_legacy_percent_column_alias() -> None:
    df = pd.DataFrame(
        {
            "Data Cabe?alho": ["2026-01-01"],
            "Equipamento": ["EQ-01"],
            "Agrup Equipamento": ["COLHEITA"],
            "Gestor": ["Gestor 1"],
            "Escala": ["PADRAO"],
            "Entregues": [3],
            "Faltantes": [0],
            "% AderÃªncia": [100.0],
            "Status": ["Completo"],
            "TURNO A": ["OK"],
            "TURNO B": ["OK"],
            "TURNO C": ["OK"],
        }
    )

    out = build_base_de_dados(df)

    assert "% Aderencia" in out.columns
    assert out.loc[0, "% Aderencia"] == 100.0


def test_indicators_template_handles_zero_total_without_nan_or_infinity() -> None:
    indicadores = pd.DataFrame(
        {
            "Aderencia Media Global": [0.0],
            "Total Esperado": [0],
            "Total Entregue": [0],
        }
    )
    resumo_status = pd.DataFrame([{"Incompleto": 0, "Ausente": 0}])

    html = render_html({"Indicadores Gerais": indicadores, "Resumo Status": resumo_status})

    assert "Infinity" not in html
    assert "NaN" not in html


def test_normalize_turno_label_maps_unknown_to_outros() -> None:
    assert normalize_turno_label("turno adm") == "TURNO A"
    assert normalize_turno_label("turno z") == "OUTROS"


def test_pivot_turnos_merges_case_insensitive_for_equipamento() -> None:
    base = pd.DataFrame(
        {
            "Data Cabecalho": [pd.Timestamp("2026-01-01")],
            "Equipamento": ["EQ-01"],
        }
    )
    imported = pd.DataFrame(
        {
            "Data Cabecalho": [pd.Timestamp("2026-01-01")],
            "Equipamento": ["eq-01"],
            "Desc.Turno": ["TURNO A"],
        }
    )

    out = pivot_turnos(imported, base)

    assert out.loc[0, "TURNO A"] == "OK"


def test_grafico_aderencia_qtd_turno_respects_adm_expected_turnos() -> None:
    df = pd.DataFrame(
        {
            "Escala": ["ADM", "PADRAO"],
            "TURNO A": ["OK", "OK"],
            "TURNO B": ["-", "-"],
            "TURNO C": ["-", "OK"],
        }
    )

    fig = grafico_aderencia_qtd_turno(df)
    ax = fig.axes[0]
    bars = [patch.get_height() for patch in ax.patches]

    # Primeiras 3 barras = entregues (A,B,C), proximas 3 = nao entregues.
    assert bars[0:3] == [2, 0, 1]
    assert bars[3:6] == [0, 1, 0]


def test_dedup_supervisores_by_agrup_keeps_one_row_per_group() -> None:
    df = pd.DataFrame(
        {
            "Gestor": ["Gestor A", "Gestor B", "Gestor C"],
            "Agrup Equipamento": ["COLHEITA", "COLHEITA", "PLANTIO"],
        }
    )

    out = _dedup_supervisores_by_agrup(df)

    assert len(out) == 2
    assert sorted(out["Agrup Equipamento"].tolist()) == ["COLHEITA", "PLANTIO"]


def test_excel_export_neutralizes_formula_like_text(tmp_path) -> None:
    final = pd.DataFrame(
        {
            "Data Cabecalho": ["2026-01-01"] * 4,
            "Equipamento": ["=1+1", "+SUM(1,1)", "-2+3", "@SUM(1,1)"],
        }
    )
    path = tmp_path / "report.xlsx"

    exportar_excel(final, {}, str(path))

    worksheet = load_workbook(path, data_only=False)["Base Consolidada"]
    values = [worksheet.cell(row=row, column=2).value for row in range(2, 6)]
    assert values == ["'=1+1", "'+SUM(1,1)", "'-2+3", "'@SUM(1,1)"]
    assert all(worksheet.cell(row=row, column=2).data_type != "f" for row in range(2, 6))


def test_email_template_contains_no_merge_conflict_markers() -> None:
    html = montar_html("01/01/2026 a 31/01/2026", "01/02/2026 10:00")

    assert "<<<<<<<" not in html
    assert "=======" not in html
    assert ">>>>>>>" not in html
    assert "GATec" not in html
    assert "Bases operacionais e de equipamentos" in html


def test_power_bi_export_neutralizes_formula_like_text(tmp_path) -> None:
    final = pd.DataFrame(
        {
            "Data Cabecalho": ["2026-01-01"],
            "Equipamento": ["=1+1"],
            "Agrup Equipamento": ["+SUM(1,1)"],
            "Gestor": ["@SUM(1,1)"],
            "Escala": ["PADRAO"],
            "Entregues": [0],
            "Faltantes": [3],
            "Aderencia": [0.0],
            "Status": ["Ausente"],
            "TURNO A": ["-"],
            "TURNO B": ["-"],
            "TURNO C": ["-"],
        }
    )

    path = atualizar_bd_excel(final, {}, out_dir=str(tmp_path))

    worksheet = load_workbook(path, data_only=False)["Base de Dados"]
    assert worksheet["B2"].value == "'=1+1"
    assert worksheet["C2"].value == "'+SUM(1,1)"
    assert worksheet["D2"].value == "'@SUM(1,1)"
    assert all(worksheet[cell].data_type != "f" for cell in ("B2", "C2", "D2"))
