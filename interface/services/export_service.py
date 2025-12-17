
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def _formatar_ws(ws):
    """Aplica AutoFilter, Freeze Panes, e ajusta larguras de coluna."""
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    # Ajuste de largura com limite para evitar colunas gigantes
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                parts = str(cell.value).split('\n')
                max_len = max(max_len, max(len(p) for p in parts))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

def _formatar_percentuais(ws):
    """
    Formata colunas percentuais:
    - Se o header contém '%', aplica formato '0.00%'.
    - Se valor estiver > 1, divide por 100 (padrão Excel).
    """
    header_values = [ws.cell(row=1, column=c).value or '' for c in range(1, ws.max_column+1)]
    for r in range(2, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            head = str(header_values[c-1])
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)) and '%' in head:
                val = cell.value
                # Ajuste robusto: se houver valores entre 0-100, converte para 0-1
                if val is not None:
                    cell.value = (val / 100.0) if val > 1 else val  # <- aqui estava &gt; no seu texto
                cell.number_format = '0.00%'

def exportar_excel(final_df: pd.DataFrame, resumos: dict, caminho: str):
    """
    Versão melhorada da sua função original:
    - Converte 'Data Cabeçalho' para dd/MM/yyyy.
    - Exporta 'Base Consolidada' e cada resumo em abas próprias.
    - Aplica filtros, congelamento, auto-largura, e formatação de % por header.
    """
    df_out = final_df.copy()
    if 'Data Cabeçalho' in df_out.columns:
        df_out['Data Cabeçalho'] = pd.to_datetime(df_out['Data Cabeçalho'], errors='coerce').dt.strftime('%d/%m/%Y')

    with pd.ExcelWriter(caminho, engine='openpyxl') as w:
        # Base Consolidada
        df_out.to_excel(w, sheet_name='Base Consolidada', index=False)

        # Resumos (garantir DataFrame)
        for nome, df_resumo in resumos.items():
            df_sheet = df_resumo
            # Se vier um Series (ex.: groupby.mean()), transforma para DataFrame
            if hasattr(df_sheet, 'to_frame') and not hasattr(df_sheet, 'columns'):
                df_sheet = df_sheet.to_frame(name='Valor')
            if not isinstance(df_sheet, pd.DataFrame):
                df_sheet = pd.DataFrame(df_sheet)
            df_sheet.to_excel(w, sheet_name=str(nome), index=True if df_sheet.index.names is not None else False)

    wb = load_workbook(caminho)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        _formatar_ws(ws)
        _formatar_percentuais(ws)
   
